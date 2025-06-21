from openpyxl import load_workbook

workbook = load_workbook("путь к xlsx файлу")

sheet = workbook.active

print(sheet.cell(row=1, column=2).value)