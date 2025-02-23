import zipfile, csv
from openpyxl.reader.excel import load_workbook
from pypdf import PdfReader
from conftest import ARCHIVE_FILE_PATH

def test_pdf_content():
    with zipfile.ZipFile(ARCHIVE_FILE_PATH) as zip_pdf:
        with zip_pdf.open('sample_pdf.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            #print(reader.pages)
            #print(len(reader.pages))
            #print(reader.pages[0].extract_text())
            assert "ПУТЕВОДИТЕЛЬ" in reader.pages[0].extract_text()
            assert len(reader.pages) == 20


def test_csv_content():
    with zipfile.ZipFile(ARCHIVE_FILE_PATH) as zip_csv:
        with zip_csv.open('sample_csv.csv') as csv_file:
            content = csv_file.read().decode('utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))
            seventh_row = csvreader[6]
            assert seventh_row[1] == 'Level 1'
            assert seventh_row[2] == '99999'


def test_xlsx_content():
    with zipfile.ZipFile(ARCHIVE_FILE_PATH) as zip_xlsx:
        with zip_xlsx.open("sample_xlsx.xlsx") as excel_file:
            workbook = load_workbook(excel_file)
            sheet = workbook.active

            content = sheet.cell(row=6, column=3).value

            assert "Техническая поддержка" in content